"""
Excel问答对加载器模块
=====================

本模块专门用于加载Excel格式的问答对数据。

Excel文件格式要求：
┌─────────────┬─────────────────────────────────────────────────────┐
│     A列     │                      B列                           │
├─────────────┼─────────────────────────────────────────────────────┤
│    问/问    │                      答/答案                         │
├─────────────┼─────────────────────────────────────────────────────┤
│  林黛玉是... │  林黛玉是《红楼梦》中的女主角...                       │
│  孙悟空的...  │  孙悟空是《西游记》中的主角...                        │
│  ...       │                      ...                           │
└─────────────┴─────────────────────────────────────────────────────┘

列名识别规则：
- 问题列：包含"问"、"question"（不区分大小写）
- 答案列：包含"答"、"answer"（不区分大小写）

返回数据格式：
```python
[
    {
        "question": "林黛玉是谁？",
        "answer": "林黛玉是《红楼梦》中的女主角...",
        "source": "/path/to/file.xlsx"
    },
    ...
]
```

使用示例：
```python
from fgcnrag.fgcn.loader import ExcelLoader

# 加载问答对
qa_pairs = ExcelLoader.load_qa_pairs("data/qa.xlsx")

for qa in qa_pairs:
    print(f"问: {qa['question']}")
    print(f"答: {qa['answer']}")
```
"""
from typing import List, Dict, Any
from pathlib import Path
import openpyxl
from langchain_core.document_loaders import BaseLoader
from langchain_core.documents import Document


class ExcelQALoader(BaseLoader):
    """
    Excel问答对加载器类

    继承LangChain的BaseLoader接口，实现统一的load方法

    功能：
    - 读取Excel文件的问答对数据
    - 自动识别问题列和答案列
    - 返回LangChain Document格式
    """

    def __init__(self, file_path: str, sheet_name: str = None):
        """
        初始化加载器

        Args:
            file_path: Excel文件的路径
            sheet_name: 工作表名称，None表示使用活动工作表
        """
        self.file_path = Path(file_path)  # 转换为Path对象
        self.sheet_name = sheet_name       # 工作表名称

    def load(self) -> List[Document]:
        """
        加载Excel中的问答对数据

        工作流程：
        1. 打开Excel文件
        2. 读取表头，定位问题列和答案列
        3. 遍历数据行，构建Document对象

        Returns:
            List[Document]: 问答对Document列表
                           如果加载失败，返回空列表
        """
        docs = []

        try:
            # ============ 打开Excel文件 ============
            #
            # openpyxl是Python操作Excel的库
            # load_workbook打开指定文件，返回工作簿对象
            wb = openpyxl.load_workbook(self.file_path)

            # 选择工作表
            # - 如果指定了sheet_name，使用该工作表
            # - 否则使用活动工作表（当前显示的工作表）
            ws = wb.active if not self.sheet_name else wb[self.sheet_name]

            # ============ 读取表头 ============
            #
            # ws[1]获取第一行（表头行）
            # 提取每列的列名
            headers = [cell.value for cell in ws[1]]

            # 初始化列索引
            question_idx = None  # 问题列索引
            answer_idx = None    # 答案列索引

            # 遍历表头，查找问题和答案列
            # 支持多种列名格式：
            # - 问、问题、question（不区分大小写）
            # - 答、答案、answer（不区分大小写）
            for idx, header in enumerate(headers):
                if header and ('问' in str(header) or 'question' in str(header).lower()):
                    question_idx = idx
                if header and ('答' in str(header) or 'answer' in str(header).lower()):
                    answer_idx = idx

            # ============ 遍历数据行 ============
            #
            # ws.iter_rows遍历所有行
            # min_row=2: 从第2行开始（跳过表头）
            # values_only=True: 返回单元格值而非Cell对象
            # start=2: 行号从2开始计数
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):

                # 确保问题列和答案列都存在
                if question_idx is not None and answer_idx is not None:

                    # 安全获取单元格值
                    # 检查索引是否越界
                    question = row[question_idx] if question_idx < len(row) else None
                    answer = row[answer_idx] if answer_idx < len(row) else None

                    # 只处理有内容的数据行
                    # 跳过空行或只有一边的数据
                    if question and answer:

                        # ============ 构建Document对象 ============
                        #
                        # page_content: 格式化的问答文本
                        #   "问题: xxx\n答案: xxx"
                        # metadata: 元数据信息
                        #   - source: 文件路径
                        #   - row: Excel行号
                        #   - question: 问题文本
                        #   - answer: 答案文本
                        doc = Document(
                            page_content=f"问题: {question}\n答案: {answer}",
                            metadata={
                                "source": str(self.file_path),    # 文件来源路径
                                "row": row_idx,                   # Excel行号
                                "question": str(question),        # 问题文本（转字符串）
                                "answer": str(answer)              # 答案文本（转字符串）
                            }
                        )
                        docs.append(doc)

            # 关闭工作簿，释放资源
            wb.close()

        except Exception as e:
            # 异常处理：打印错误并返回已加载的文档
            print(f"加载Excel失败: {e}")

        return docs


class ExcelLoader:
    """
    Excel工具类

    提供便捷的问答对加载方法

    内部使用ExcelQALoader，但返回更简洁的字典格式
    """

    @staticmethod
    def load_qa_pairs(file_path: str) -> List[Dict[str, Any]]:
        """
        加载问答对数据

        将ExcelQALoader返回的Document格式转换为字典列表格式
        更适合后续处理和向量化

        Args:
            file_path: Excel文件的路径

        Returns:
            List[Dict]: 问答对列表
                       每项包含：
                       - question: 问题文本
                       - answer: 答案文本
                       - source: 文件来源路径
        """
        # 创建加载器实例
        loader = ExcelQALoader(file_path)

        # 加载文档
        docs = loader.load()

        # 转换为字典格式
        return [
            {
                "question": doc.metadata.get("question", ""),   # 从metadata提取问题
                "answer": doc.metadata.get("answer", ""),       # 从metadata提取答案
                "source": doc.metadata.get("source", "")         # 从metadata提取来源
            }
            for doc in docs  # 遍历所有Document
        ]
